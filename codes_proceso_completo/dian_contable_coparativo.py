import os
from openpyxl import load_workbook, Workbook
import re
import pandas as pd

# Función para cargar un archivo .xlsx
def cargar_archivo(ruta):
    if os.path.exists(ruta):
        return ruta
    raise ValueError(f"No se ha encontrado el archivo en la ruta especificada: {ruta}")

UPLOAD_FOLDER = os.path.abspath("")

# Obtener la lista de subcarpetas dentro de "archivos_usuarios"
subcarpetas = [os.path.join(UPLOAD_FOLDER, "archivos_usuarios", d) for d in os.listdir(os.path.join(UPLOAD_FOLDER, "archivos_usuarios")) if os.path.isdir(os.path.join(UPLOAD_FOLDER, "archivos_usuarios", d))]

# Asegurarse de que hay subcarpetas disponibles
if not subcarpetas:
    raise ValueError("No se encontraron subcarpetas dentro de la carpeta 'archivos_usuarios'.")

# Iterar sobre cada subcarpeta para procesar los archivos dentro de ella
for subcarpeta in subcarpetas:
    ruta_archivo_dian = os.path.join(subcarpeta, 'DIAN.xlsx')
    ruta_archivo_sinco = os.path.join(subcarpeta, 'SINCO.xlsx')
    ruta_archivo_salida = os.path.join(subcarpeta, 'archivo final.xlsx')

    # Procesar el archivo DIAN
    df_dian = pd.read_excel(cargar_archivo(ruta_archivo_dian))
    df_dian['Prefijo'] = df_dian['Prefijo'].fillna('')
    df_dian['Folio'] = df_dian['Folio'].fillna('')
    df_dian['Documento del Tercero'] = df_dian['Prefijo'].astype(str) + df_dian['Folio'].astype(str)
    df_dian['Subtotal'] = df_dian['Total'].fillna(0) - df_dian['IVA'].fillna(0) - df_dian['INC'].fillna(0)

    df_dian = df_dian.drop(columns=['Prefijo', 'Folio', 'NIT Receptor', 'Nombre Receptor'])
    columnas_reordenadas_dian = ['Tipo de documento', 'CUFE/CUDE', '', 'Fecha Emisión', 'Fecha Recepción', 'NIT Emisor', 'Nombre Emisor', 'IVA', 'ICA', 'INC','Subtotal','Total', 'Estado', 'Grupo']
    df_dian = df_dian[columnas_reordenadas_dian]
    nombre_archivo_modificado_dian =  os.path.join(subcarpeta, "5.2. RELACION DIAN_modificado.xlsx")
    df_dian.to_excel(nombre_archivo_modificado_dian, index=False)

    # Procesar el archivo SINCO
    wb = load_workbook(cargar_archivo(ruta_archivo_sinco))
    hoja = wb.active
    nuevo_wb = Workbook()
    nueva_hoja = nuevo_wb.active
    encabezados_encontrados = False
    encabezados = []

    for fila in hoja.iter_rows(values_only=True):
        if fila[0] == 'Tipo Registro':
            encabezados_encontrados = True
            encabezados = list(fila)
            tipo_doc_index = encabezados.index('Tipo Doc.')
            consecutivo_index = encabezados.index('Consecutivo')
            encabezados[tipo_doc_index] = 'Doc. Contable'
            del encabezados[consecutivo_index]
            nueva_hoja.append(encabezados)
            continue
        if encabezados_encontrados:
            cuenta_contable = fila[1]
            if not cuenta_contable.startswith(('539595')):
                nueva_fila = list(fila)
                doc_contable = f"{nueva_fila[tipo_doc_index]} {nueva_fila[consecutivo_index]}"
                nueva_fila[tipo_doc_index] = doc_contable
                del nueva_fila[consecutivo_index]
                documento_tercero_index = encabezados.index('Documento del Tercero')
                documento_tercero = nueva_fila[documento_tercero_index]
                if isinstance(documento_tercero, str):
                    documento_tercero = re.sub(r'[-._ ]', '', documento_tercero)
                else:
                    documento_tercero = str(documento_tercero) if documento_tercero is not None else ''
                nueva_fila[documento_tercero_index] = documento_tercero
                nueva_hoja.append(nueva_fila)

    # Guardar el archivo "MovDocCuenta_filtrado.xlsx" en la subcarpeta
    nuevo_nombre_archivo_sinco = os.path.join(subcarpeta, "MovDocCuenta_filtrado.xlsx")
    nuevo_wb.save(nuevo_nombre_archivo_sinco)

    # Procesar el archivo generado de SINCO
    df_sinco = pd.read_excel(nuevo_nombre_archivo_sinco)
    df_sinco[['Numero de Factura', 'Concepto Restante']] = df_sinco['Concepto'].str.extract(r'([A-Z0-9\-]+)\s+([A-Z].*)')
    df_sinco['Numero de Factura'] = df_sinco['Numero de Factura'].str.replace(r'[-_.]', '', regex=True)
    df_sinco.drop(columns=['Concepto'], inplace=True)

    # Guardar el archivo "MovDocCuenta_tratado.xlsx" en la subcarpeta
    nombre_archivo_tratado_sinco = os.path.join(subcarpeta, "MovDocCuenta_tratado.xlsx")
    df_sinco.to_excel(nombre_archivo_tratado_sinco, index=False)

    # Fusionar y comparar los archivos
    archivo1 = nombre_archivo_modificado_dian
    archivo2 = nombre_archivo_tratado_sinco
    archivo_salida = ruta_archivo_salida

    # Leer los archivos Excel
    df1 = pd.read_excel(archivo1, sheet_name='Sheet1')  # Lee la hoja 'Sheet1'
    df2 = pd.read_excel(archivo2, sheet_name='Sheet1')  # Lee la hoja 'Sheet1'

    # Obtener las columnas relevantes
    documentos_tercero = df1['Documento del Tercero']
    nit_emisor = df1['NIT Emisor']
    numeros_factura = df2['Numero de Factura']
    nit_receptor = df2['NIT']
    fechas_contabilizacion = df2['Fecha']
    docs_contables = df2['Doc. Contable']
    documentos_tercero_2 = df2['Documento del Tercero']

    # Crear una lista de valores para las nuevas columnas "Estado de Factura" y "Fecha Contabilización"
    estado_factura = []
    fechas_contabilizacion_nueva = []
    docs_contables_nuevos = []

    for doc, nit in zip(documentos_tercero, nit_emisor):
        # Verificar coincidencia en "Numero de Factura" y "Documento del Tercero"
        index_factura = (numeros_factura == doc) & (nit_receptor == nit)
        index_doc_tercero = (documentos_tercero_2 == doc) & (nit_receptor == nit)

        if any(index_factura) or any(index_doc_tercero):
            estado_factura.append('Contabilizado')
            if any(index_factura):
                fechas_contabilizacion_nueva.append(fechas_contabilizacion[index_factura].iloc[0] if index_factura.any() else None)
                docs_contables_nuevos.append(docs_contables[index_factura].iloc[0] if index_factura.any() else None)
            else:
                fechas_contabilizacion_nueva.append(fechas_contabilizacion[index_doc_tercero].iloc[0] if index_doc_tercero.any() else None)
                docs_contables_nuevos.append(docs_contables[index_doc_tercero].iloc[0] if index_doc_tercero.any() else None)
        else:
            estado_factura.append('Revisar')
            fechas_contabilizacion_nueva.append(None)
            docs_contables_nuevos.append(None)

    # Agregar las columnas "Estado de Factura", "Fecha Contabilización" y "Documento Contable" al DataFrame df1
    df1['Estado de Factura'] = estado_factura
    df1['Fecha Contabilización'] = fechas_contabilizacion_nueva
    df1['Documento Contable'] = docs_contables_nuevos

    # Guardar el DataFrame modificado en un nuevo archivo Excel
    df1.to_excel(archivo_salida, index=False)

    # Eliminar archivos temporales, excluyendo "MovDocCuenta_tratado.xlsx"
    archivos_temporales = [nombre_archivo_modificado_dian, nuevo_nombre_archivo_sinco]
    for archivo in archivos_temporales:
        try:
            os.remove(archivo)
        except Exception as e:
            pass  # Si hay algún error al eliminar el archivo, continuar sin mostrar mensaje

print("Código ejecutado exitosamente")
