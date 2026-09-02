import os
import zipfile
import fitz  # PyMuPDF
import shutil
import openpyxl
import sys

# Aceptar ruta de carpeta y timestamp específico como argumentos
if len(sys.argv) > 1:
    folder_path = sys.argv[1]
else:
    UPLOAD_FOLDER = os.path.abspath("")
    folder_path = os.path.join(UPLOAD_FOLDER, "archivos_usuarios")

timestamp_filter = sys.argv[2] if len(sys.argv) > 2 else None

# Lista de archivos a excluir
excluir_archivos = [
    'Cuenta_contable.xlsx',
    'MovDocCuenta_tratado.xlsx',
    'SINCO.xlsx',
    'DIAN.xlsx',
    'MovDocCuenta_Excel.xlsx'
]

# Función para buscar variables en el texto extraído del documento PDF
def buscar_variables(documento):
    variables = {
        "Número de Factura:": None,
        "Fecha de Emisión:": None,
        "Razón Social:": None
    }

    for page_num in range(documento.page_count):
        page = documento.load_page(page_num)
        texto = page.get_text()

        for variable in variables:
            if variable in texto:
                index = texto.find(variable)
                if index != -1:
                    start_index = index + len(variable)
                    end_index = texto.find('\n', start_index)
                    if end_index == -1:
                        end_index = None
                    variables[variable] = texto[start_index:end_index].strip()

    return variables

# Función para aplicar el filtro y la inmovilización de la primera fila en un archivo Excel (.xlsx)
def aplicar_filtro_y_inmovilizar_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        if worksheet.max_row > 1 and worksheet.max_column > 1:
            worksheet.auto_filter.ref = worksheet.dimensions  # Aplicar el filtro a la primera fila
            worksheet.freeze_panes = worksheet['A2']  # Inmovilizar la primera fila (todo lo anterior a A2)
    workbook.save(file_path)

# Determinar qué subcarpetas procesar
if timestamp_filter:
    subcarpetas_a_procesar = [timestamp_filter]
else:
    subcarpetas_a_procesar = [d for d in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, d))]

# Recorrer las subcarpetas a procesar
for subfolder in subcarpetas_a_procesar:
    subfolder_path = os.path.join(folder_path, subfolder)
    
    # Verificar que sea una subcarpeta
    if not os.path.isdir(subfolder_path):
        continue
    
    try:
        # Ruta del archivo zip que se va a crear con el mismo nombre de la subcarpeta
        zip_file_path = os.path.join(folder_path, f'{subfolder}.zip')
        
        with zipfile.ZipFile(zip_file_path, 'w') as zipf:
            # Recopilar y ordenar los archivos (.xls, .xlsx, luego .pdf)
            archivos_xls = []
            archivos_xlsx = []
            archivos_pdf = []

            for foldername, subfolders, filenames in os.walk(subfolder_path):
                for filename in filenames:
                    if filename in excluir_archivos:
                        continue  # Omitir los archivos que están en la lista de exclusión

                    file_path = os.path.join(foldername, filename)
                    if filename.endswith('.xls'):
                        archivos_xls.append((file_path, os.path.relpath(file_path, subfolder_path)))
                    elif filename.endswith('.xlsx'):
                        archivos_xlsx.append((file_path, os.path.relpath(file_path, subfolder_path)))
                    elif filename.endswith('.pdf'):
                        archivos_pdf.append((file_path, os.path.relpath(file_path, subfolder_path)))

            # Procesar archivos .xls sin modificar
            for file_path, arcname in archivos_xls:
                zipf.write(file_path, arcname)

            # Procesar archivos .xlsx (aplicar filtro e inmovilizar)
            for file_path, arcname in archivos_xlsx:
                aplicar_filtro_y_inmovilizar_xlsx(file_path)  # Aplicar el filtro y la inmovilización antes de añadir al ZIP
                zipf.write(file_path, arcname)

            # Añadir los archivos PDF renombrados
            for file_path, arcname in archivos_pdf:
                # Abrir el archivo PDF
                documento = fitz.open(file_path)

                # Buscar las variables en el documento
                variables_encontradas = buscar_variables(documento)

                # Cerrar el documento
                documento.close()

                # Construir el nuevo nombre de archivo
                numero_factura = variables_encontradas.get('Número de Factura:') or 'SIN_NUMERO'
                razon_social = variables_encontradas.get('Razón Social:') or 'SIN_RAZON'
                nuevo_nombre = f"{numero_factura}_{razon_social}.pdf"
                nuevo_nombre = nuevo_nombre.replace("/", "-")  # Reemplazar '/' por '-' para evitar problemas en el nombre del archivo

                # Añadir el archivo PDF al zip con el nuevo nombre
                zipf.write(file_path, os.path.join(os.path.dirname(arcname), nuevo_nombre))

        # Eliminar la subcarpeta original
        shutil.rmtree(subfolder_path)

    except Exception as e:
        # Si se creó un zip parcial, eliminarlo
        zip_file_path_parcial = os.path.join(folder_path, f'{subfolder}.zip')
        if os.path.exists(zip_file_path_parcial):
            try:
                os.remove(zip_file_path_parcial)
            except:
                pass
